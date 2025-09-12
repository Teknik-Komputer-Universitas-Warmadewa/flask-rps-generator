from flask import Flask, render_template, request, send_file, redirect, url_for, abort, send_from_directory
import openpyxl
import os
import io
import xlsxwriter
from werkzeug.utils import secure_filename
from datetime import datetime
from collections import defaultdict
from xlsxwriter.utility import xl_rowcol_to_cell
import re
import logging
from logging.handlers import TimedRotatingFileHandler
# import string

app = Flask(__name__)
# Get the absolute path of the current directory
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# Setup logging
LOG_FILE = os.path.join(BASE_DIR, "rps_generator.log")

logger = logging.getLogger("RPSGenerator")

# Folder untuk menyimpan data upload
UPLOAD_FOLDER = os.path.join(BASE_DIR, "uploads")
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER

# Path ke Excel Template daftar matkul
EXCEL_FILE = os.path.join(BASE_DIR, "data", "Final Template Kurikulum 2025.xlsx")

SHEET_MATKUL = "9. Susunan Mata Kuliah"
SHEET_CPL = "2. CPL Prodi"
SHEET_CPMK = "12.2. list CPMK"
SHEET_SUBCPMK = "15. Pemetaan MK-CPMK-Su"

MATKUL_START_ROW = 3
MATKUL_END_ROW = 69
MATKUL_COLUMN = "D"


def get_matkul_list():
    """Read matkul list from Excel file"""
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
        sheet = wb[SHEET_MATKUL]
        matkul = [
            sheet[f"{MATKUL_COLUMN}{row}"].value
            for row in range(MATKUL_START_ROW, MATKUL_END_ROW + 1)
            if sheet[f"{MATKUL_COLUMN}{row}"].value is not None
        ]
        wb.close()
        return matkul
    except Exception as e:
        # Log error agar tahu penyebabnya, tapi program tetap jalan
        print(f"[WARNING] Gagal membaca file Excel List MK: {EXCEL_FILE} -> {e}")

def get_rps_data(nama_matkul):
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
        sheet = wb[SHEET_MATKUL]

        result = {
            "kode_matkul": None,
            "semester": None,
            "rumpun": None,
            "bobot_sks": None
        }

        for row in range(MATKUL_START_ROW, MATKUL_END_ROW + 1):
            matkul_name = sheet[f"D{row}"].value  # cari nama matkul di kolom D
            if matkul_name and str(matkul_name).strip().lower() == str(nama_matkul).strip().lower():
                result["kode_matkul"] = str(int(sheet[f"C{row}"].value))
                result["semester"] = sheet[f"N{row}"].value
                result["rumpun"] = sheet[f"O{row}"].value
                result["bobot_sks"] = sheet[f"E{row}"].value
                break  # berhenti setelah ketemu

        wb.close()
        return result
    except Exception as e:
        # Log error agar tahu penyebabnya, tapi program tetap jalan
        print(f"[WARNING] Gagal membaca file Excel RPS data: {EXCEL_FILE} -> {e}")

def get_cpl_cpmk_sub_list(nama_matkul):
    """Ambil daftar CPL, CPMK, dan SubCPMK berdasarkan nama matkul"""
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
        sheet = wb[SHEET_SUBCPMK]

        cpls_kode, cpls_desc = [], []
        cpmks_kode, cpmks_desc = [], []
        subcpmks_kode, subcpmks_desc = [], []

        for row in range(3, 273):  # B3:F272 + H:I
            mk = sheet[f"B{row}"].value
            if mk and mk.strip().lower() == nama_matkul.strip().lower():
                # ambil CPL
                cpl_kode = sheet[f"H{row}"].value
                cpl_desc = sheet[f"I{row}"].value
                if cpl_kode: cpls_kode.append(str(cpl_kode))
                if cpl_desc: cpls_desc.append(str(cpl_desc))

                # ambil CPMK
                cpmk_kode = sheet[f"C{row}"].value
                cpmk_desc = sheet[f"D{row}"].value
                if cpmk_kode: cpmks_kode.append(str(cpmk_kode))
                if cpmk_desc: cpmks_desc.append(str(cpmk_desc))

                # ambil SubCPMK
                subcpmk_kode = sheet[f"E{row}"].value
                subcpmk_desc = sheet[f"F{row}"].value
                if subcpmk_kode: subcpmks_kode.append(str(subcpmk_kode))
                if subcpmk_desc: subcpmks_desc.append(str(subcpmk_desc))

        wb.close()
        return {
            "cpl_kode": cpls_kode,
            "cpl_desc": cpls_desc,
            "cpmk_kode": cpmks_kode,
            "cpmk_desc": cpmks_desc,
            "subcpmk_kode": subcpmks_kode,
            "subcpmk_desc": subcpmks_desc,
        }
    except Exception as e:
        # Log error agar tahu penyebabnya, tapi program tetap jalan
        print(f"[WARNING] Gagal membaca file Excel Sub CPMK: {EXCEL_FILE} -> {e}")

def get_matkul_data(nama_matkul, tahun):
    """Ambil semua data terkait matkul dari file data_[matkul]_[tahun].xlsx"""
    # filename = f"uploads/data_{nama_matkul}_{tahun}.xlsx"
    # wb = openpyxl.load_workbook(filename, data_only=True)

    # Use absolute path
    filename = os.path.join(UPLOAD_FOLDER, f"data_{nama_matkul}_{tahun}.xlsx")
    
    try:
        wb = openpyxl.load_workbook(filename, data_only=True)
    except FileNotFoundError:
        raise ValueError(f"File '{filename}' tidak ditemukan")
    except Exception as e:
        raise ValueError(f"Error membuka file '{filename}': {str(e)}")
    
    sheet_name = nama_matkul.split()[0]
    if sheet_name not in wb.sheetnames:
        wb.close()
        raise ValueError(f"Sheet '{nama_matkul}' tidak ditemukan dalam {filename}")

    sheet = wb[sheet_name]

    # pustaka, tim, syarat
    pustaka_utama, pustaka_pendukung, team_teaching, nik, matkul_syarat = [], [], [], [], []

    # Pertemuan
    minggu_ke, subcpmk_weekly, indikator, kriteria, kriteria_numbered, materi, bobot, pustaka_weekly = [], [], [], [], [], [], [], []
    materi_non_uts_uas, materi_non_uts_uas_numbered, materi_weekly_numbered = [], [], []

    # Kelas
    kelas, jml_mhs, hari, tempat, tahun_ajar = [], [], [], [], []

    # CPL/CPMK/subCPMK bobot
    cpl_bobot, cpmk_bobot, subcpmk_bobot, total_bobot = [], [], [], []

    # iterasi baris mulai baris ke-2
    for row in sheet.iter_rows(min_row=2, values_only=True):
        # A:E
        col_a, col_b, col_c, col_d, col_e = row[0:5]
        # G:M (kolom 6-11, total 6 kolom)
        col_g, col_h, col_i, col_j, col_k, col_l, col_m = row[6:13]
        # O:Q (kolom 14-16, total 3 kolom)
        col_o, col_p, col_q = row[14:17]
        # Y
        col_y = row[24]
        # AA:AE
        col_aa, col_ab, col_ac, col_ad, col_ae = row[26:31]

        if col_a: pustaka_utama.append(str(col_a))
        if col_b: pustaka_pendukung.append(str(col_b))
        if col_c: team_teaching.append(str(col_c))
        if col_d: nik.append(str(col_d))
        if col_e: matkul_syarat.append(str(col_e))

        if col_g: minggu_ke.append(str(col_g))
        if col_h: subcpmk_weekly.append(str(col_h))
        if col_i: indikator.append(str(col_i))
        if col_j: kriteria.append(str(col_j))
        if col_k: materi.append(str(col_k))
        if col_l: bobot.append(str(col_l))
        if col_m: pustaka_weekly.append(str(col_m))

        if col_o: cpl_bobot.append(str(col_o))
        if col_p: cpmk_bobot.append(str(col_p))
        if col_q: subcpmk_bobot.append(str(col_q))
        if col_y: total_bobot.append(str(col_y))

        if col_aa: kelas.append(str(col_aa))
        if col_ab: jml_mhs.append(str(col_ab))
        if col_ac: hari.append(str(col_ac))
        if col_ad: tempat.append(str(col_ad))
        if col_ae: tahun_ajar.append(str(col_ae))

    wb.close()

    # --- Olahan materi ---
    exclude_keywords = ["Evaluasi UTS", "Evaluasi UAS", "Proyek Akhir"]

    # Filter materi, buang yang mengandung kata di exclude_keywords
    materi_non_uts_uas = [m for m in materi if not any(kw in str(m) for kw in exclude_keywords)]

    # Tambahkan numbering untuk materi non-UTS/UAS
    materi_non_uts_uas_numbered = [f"{i+1}. {m}" for i, m in enumerate(materi_non_uts_uas)]

    # Numbering untuk semua materi, tapi jangan loncati index
    materi_weekly_numbered = []
    counter = 1
    for m in materi:
        if any(kw in str(m) for kw in exclude_keywords):
            materi_weekly_numbered.append(m)  # tampilkan apa adanya, tanpa nomor
        else:
            materi_weekly_numbered.append(f"{counter}. {m}")
            counter += 1  # hanya naik kalau materi bukan evaluasi

    # --- Olahan kriteria (penomoran per jenis) ---
    kriteria_numbered = []
    counter_map = {}  # simpan hitungan per jenis
    for k in kriteria:
        if not k: 
            continue
        if "Evaluasi UTS" in k or "Evaluasi UAS" in k:
            kriteria_numbered.append(k)
            continue
        if ":" in k:
            jenis, isi = k.split(":", 1)
            jenis = jenis.strip()
            counter_map[jenis] = counter_map.get(jenis, 0) + 1
            kriteria_numbered.append(f"{jenis} {counter_map[jenis]}:{isi.strip()}")
        else:
            kriteria_numbered.append(k)

    # --- Rubrik ---
    def extract_rubrik(tag):
        """Cari subcpmk dari weekly jika kriteria mengandung [tag]"""
        return list({subcpmk_weekly[i] for i, k in enumerate(kriteria) if k and f"[{tag}]" in k})

    rubrik_SP1_subcpmk = extract_rubrik("SP1")
    rubrik_H1_subcpmk = extract_rubrik("H1")
    rubrik_H2_subcpmk = extract_rubrik("H2")
    rubrik_H3_subcpmk = extract_rubrik("H3")
    rubrik_A1_subcpmk = extract_rubrik("A1")
    rubrik_A2_subcpmk = extract_rubrik("A2")
    rubrik_A3_subcpmk = extract_rubrik("A3")

    # fungsi lookup subcpmk → CPL
    def map_to_cpl(subcpmk_list):
        mapped = []
        for sc in subcpmk_list:
            if sc in subcpmk_bobot:
                idx = subcpmk_bobot.index(sc)
                if idx < len(cpl_bobot):
                    mapped.append(cpl_bobot[idx])
        return mapped

    rubrik_SP1_cpl = map_to_cpl(rubrik_SP1_subcpmk)
    rubrik_H1_cpl = map_to_cpl(rubrik_H1_subcpmk)
    rubrik_H2_cpl = map_to_cpl(rubrik_H2_subcpmk)
    rubrik_H3_cpl = map_to_cpl(rubrik_H3_subcpmk)
    rubrik_A1_cpl = map_to_cpl(rubrik_A1_subcpmk)
    rubrik_A2_cpl = map_to_cpl(rubrik_A2_subcpmk)
    rubrik_A3_cpl = map_to_cpl(rubrik_A3_subcpmk)

    def nomor_indikator(subcpmk_weekly, subcpmk_bobot, indikator):
        exclude_keywords = ["Evaluasi UTS", "Evaluasi UAS"]

        # counter untuk setiap subcpmk
        subcpmk_counter = {k: 0 for k in subcpmk_bobot}
        indikator_numbered = []

        for sub, ind in zip(subcpmk_weekly, indikator):
            if not ind:
                indikator_numbered.append("")
            elif any(kw in str(ind) for kw in exclude_keywords):
                indikator_numbered.append(ind)  # tampilkan apa adanya
            else:
                if sub in subcpmk_bobot:
                    idx = subcpmk_bobot.index(sub) + 1  # index di subcpmk_bobot (mulai dari 1)
                    subcpmk_counter[sub] += 1           # urutan keberapa dalam subcpmk
                    nomor = f"{idx}.{subcpmk_counter[sub]}"
                    indikator_numbered.append(f"{nomor} {ind}")
                else:
                    indikator_numbered.append(ind)  # fallback kalau sub tidak ada di bobot
        return indikator_numbered
    
    # --- Olahan indikator dengan nomor per subcpmk ---
    indikator_numbered = nomor_indikator(subcpmk_weekly, subcpmk_bobot, indikator)

    bobot_dict = defaultdict(int)
    for i_cpmk, i_bobot in zip(cpmk_bobot, total_bobot):
        bobot_dict[i_cpmk] += int(float(i_bobot))

    seen = []
    bobot_per_cpmk = []
    for cpmk in cpmk_bobot:
        if cpmk not in seen:
            seen.append(cpmk)
            bobot_per_cpmk.append(bobot_dict[cpmk])

    return {
        "pustaka_utama": pustaka_utama,
        "pustaka_pendukung": pustaka_pendukung,
        "team_teaching": team_teaching,
        "nik": nik,
        "matkul_syarat": matkul_syarat,
        "minggu_ke": minggu_ke,
        "subcpmk_weekly": subcpmk_weekly,
        "indikator": indikator,
        "indikator_numbered": indikator_numbered,
        "materi": materi,
        "bobot": bobot,
        "pustaka_weekly": pustaka_weekly,
        "kelas": kelas,
        "jml_mhs": jml_mhs,
        "hari": hari,
        "tempat": tempat,
        "tahun_ajar": tahun_ajar,
        "materi_non_uts_uas": materi_non_uts_uas,
        "materi_non_uts_uas_numbered": materi_non_uts_uas_numbered,
        "materi_weekly_numbered": materi_weekly_numbered,
        "kriteria": kriteria,
        "kriteria_numbered": kriteria_numbered,
        "cpl_bobot": cpl_bobot,
        "cpmk_bobot": cpmk_bobot,
        "bobot_per_cpmk": bobot_per_cpmk,
        "subcpmk_bobot": subcpmk_bobot,
        "total_bobot": total_bobot,
        "rubrik_SP1_subcpmk": rubrik_SP1_subcpmk,
        "rubrik_SP1_cpl": rubrik_SP1_cpl,
        "rubrik_H1_subcpmk": rubrik_H1_subcpmk,
        "rubrik_H1_cpl": rubrik_H1_cpl,
        "rubrik_H2_subcpmk": rubrik_H2_subcpmk,
        "rubrik_H2_cpl": rubrik_H2_cpl,
        "rubrik_H3_subcpmk": rubrik_H3_subcpmk,
        "rubrik_H3_cpl": rubrik_H3_cpl,
        "rubrik_A1_subcpmk": rubrik_A1_subcpmk,
        "rubrik_A1_cpl": rubrik_A1_cpl,
        "rubrik_A2_subcpmk": rubrik_A2_subcpmk,
        "rubrik_A2_cpl": rubrik_A2_cpl,
        "rubrik_A3_subcpmk": rubrik_A3_subcpmk,
        "rubrik_A3_cpl": rubrik_A3_cpl,
    }

@app.route("/", methods=["GET", "POST"])
def index():
    matkul_list = get_matkul_list()
    selected_matkul = None
    tahun = datetime.now().year  # default tahun sekarang
    uploaded_file = None  # <-- tambahkan default None

    if request.method == "POST":
        selected_matkul = request.form.get("nama_matkul")
        tahun = request.form.get("tahun") or str(datetime.now().year)

        # handle file upload
        if "rps_file" in request.files:
            file = request.files["rps_file"]
            if file.filename:
                safe_name = secure_filename(file.filename)
                ext = os.path.splitext(safe_name)[1]  # ambil ekstensi
                new_filename = f"data_{selected_matkul}_{tahun}{ext}"
                save_path = os.path.join(app.config["UPLOAD_FOLDER"], new_filename)
                file.save(save_path)
                uploaded_file = new_filename  # <-- simpan nama file

    return render_template(
        "index.html",
        matkul_list=matkul_list,
        selected_matkul=selected_matkul,
        tahun=tahun,
        uploaded_file=uploaded_file,  # <-- kirim ke template
    )


@app.route("/download-rps", methods=["POST"])
def download_rps():
    matkul = request.form.get("nama_matkul")
    tahun = request.form.get("tahun") or str(datetime.now().year)

    # cpl_cpmk_sub = get_cpl_cpmk_sub_list(matkul)
    # matkul_data = get_matkul_data(matkul,tahun)
    # rps_data = get_rps_data(matkul)

    try:
        # Log the attempt
        logger.info(f"Attempting to generate RPS for {matkul} ({tahun})")
        
        cpl_cpmk_sub = get_cpl_cpmk_sub_list(matkul)
        logger.info(f"Successfully retrieved CPL/CPMK/SubCPMK data")
        
        matkul_data = get_matkul_data(matkul, tahun)
        logger.info(f"Successfully retrieved matkul data")
        
        rps_data = get_rps_data(matkul)
        logger.info(f"Successfully retrieved RPS data")
        
    except FileNotFoundError as e:
        logger.error(f"File not found: {e}")
        abort(404, description=f"File data untuk mata kuliah '{matkul}' tahun {tahun} tidak ditemukan. Pastikan file sudah diupload.")
    except ValueError as e:
        logger.error(f"Data error: {e}")
        abort(400, description=str(e))
    except Exception as e:
        logger.error(f"Unexpected error: {e}")
        abort(500, description=f"Terjadi kesalahan sistem: {str(e)}")

    if not matkul:
        abort(400, description="Nama mata kuliah wajib diisi")

    
    try:
        output = io.BytesIO()
        workbook = xlsxwriter.Workbook(output, {"in_memory": True})
        worksheet = workbook.add_worksheet("RPS")

        # Atur ukuran kolom
        worksheet.set_column("A:A", 5)        # Kolom A kecil
        worksheet.set_column("B:B", 18)       # Kolom B - L agak besar (2x normal)
        worksheet.set_column("C:C", 36)       # Kolom B - L agak besar (2x normal)
        worksheet.set_column("D:L", 18)       # Kolom B - L agak besar (2x normal)

        # Format header dengan background hitam, font putih, center
        header_small = workbook.add_format({
            "align": "center",
            "valign": "vcenter",
            "font_name": "Tahoma",
            "font_size": 12,
            "border": 1,
            "bold": True,
            "font_color": "white",
            "bg_color": "black",
            "text_wrap": True
        })

        header_big = workbook.add_format({
            "font_name": "Tahoma",
            "font_size": 28,
            "border": 1,
            "bold": True,
            "align": "center",
            "font_color": "white",
            "bg_color": "black",
            "valign": "vcenter"
        })

        header_medium = workbook.add_format({
            "font_name": "Tahoma",
            "font_size": 18,
            "border": 1,
            "bold": True,
            "align": "center",
            "font_color": "white",
            "bg_color": "black",
            "valign": "vcenter"
        })

        # Format judul: bg abu + font hitam
        title_format = workbook.add_format({
            "font_name": "Tahoma",
            "font_size": 12,
            "border": 1,
            "align": "center",
            "valign": "vcenter",
            "bold": True,
            "font_color": "black",
            "text_wrap": True,
            "bg_color": "#C0C0C0"  # abu-abu
        })

        # Format judul: bg abu + font hitam
        title_format_green = workbook.add_format({
            "font_name": "Tahoma",
            "font_size": 12,
            "border": 1,
            "align": "center",
            "valign": "vcenter",
            "bold": True,
            "font_color": "black",
            "text_wrap": True,
            "bg_color": "green"  # abu-abu
        })

        title_cpl_format = workbook.add_format({
            "font_name": "Tahoma",
            "font_size": 12,
            "border": 1,
            "align": "left",
            "valign": "vcenter",
            "bold": True,
            "font_color": "black",
            "bg_color": "#C0C0C0"  # abu-abu
        })

        title_korelasi_format = workbook.add_format({
            "font_name": "Tahoma",
            "font_size": 12,
            "border": 1,
            "align": "center",
            "valign": "vcenter",
            "bold": True,
            "font_color": "black",
            "text_wrap": True
        })

        # Format text
        text_format = workbook.add_format({
            "font_name": "Tahoma",
            "font_size": 12,
            "border": 1,
            "align": "center",
            "valign": "vcenter",
            "text_wrap": True,
            "font_color": "black"
        })

        # Format text
        text_cpl_format = workbook.add_format({
            "font_name": "Tahoma",
            "font_size": 12,
            "border": 1,
            "align": "left",
            "valign": "vcenter",
            "font_color": "black",
            "text_wrap": True
        })

        # Format text
        text_otorisasi_format = workbook.add_format({
            "font_name": "Tahoma",
            "font_size": 12,
            "border": 1,
            "align": "center",
            "valign": "bottom",
            "font_color": "black"
        })

        date_format = workbook.add_format({
            "font_name": "Tahoma",
            "font_size": 12,
            "align": "center",
            "valign": "vcenter",
            "border": 1,
            "num_format": "dd-mm-yyyy"
        })

        percent_format_bold = workbook.add_format({
            "font_name": "Tahoma",
            "font_size": 12,
            "border": 1,
            "align": "center",
            "valign": "vcenter",
            "bold": True,
            "font_color": "black",
            "text_wrap": True,
            "num_format": "0%"
        })

        percent_format_bold_fill = workbook.add_format({
            "font_name": "Tahoma",
            "font_size": 12,
            "border": 1,
            "align": "center",
            "valign": "vcenter",
            "bold": True,
            "font_color": "black",
            "text_wrap": True,
            "num_format": "0%",
            "bg_color": "#C0C0C0"
        })

        percent_format = workbook.add_format({
            "font_name": "Tahoma",
            "font_size": 12,
            "border": 1,
            "align": "center",
            "valign": "vcenter",
            "font_color": "black",
            "text_wrap": True,
            "num_format": "0%"
        })

        # Header
        # Tambahkan logo
        # Set tinggi baris header (sedikit lebih tinggi dari normal)
        worksheet.set_row(1, 22)  # baris 2
        worksheet.set_row(2, 22)  # baris 3
        worksheet.set_row(3, 22)  # baris 4
        worksheet.set_row(4, 26)  # baris 5
        worksheet.merge_range("B2:B5", "", header_medium)
        worksheet.insert_image("B2", "data/logo.png", {
            "x_scale": 1.5,  # perkecil jika perlu
            "y_scale": 1.5,
            "x_offset": 10,  # sedikit geser biar rapi
            "y_offset": 2,
        })
        
        worksheet.merge_range("C2:J2", "UNIVERSITAS WARMADEWA", header_medium)
        worksheet.merge_range("C3:J3", "FAKULTAS TEKNIK DAN PERENCANAAN", header_medium)
        worksheet.merge_range("C4:J4", "PROGRAM STUDI TEKNIK KOMPUTER", header_medium)
        worksheet.merge_range("C5:J5", "RENCANA PEMBELAJARAN SEMESTER", header_big)

        kode_dokumen_rps = f'FTP-TKOM-RPS-{rps_data["kode_matkul"]}-{tahun}'
        worksheet.merge_range("K2:L3", "Kode Dokumen", header_small)
        worksheet.merge_range("K4:L5", str(kode_dokumen_rps), header_small)

        # Matakuliah Info
        worksheet.merge_range("B6:C6", "MATA KULIAH (MK)", title_format)
        worksheet.merge_range("D6:E6", "KODE", title_format)
        worksheet.write("F6", "RUMPUN MK", title_format)
        worksheet.merge_range("G6:I6", "BOBOT (SKS)", title_format)
        worksheet.write("J6", "SEMESTER", title_format)
        worksheet.merge_range("K6:L6", "Tgl. PENETAPAN", title_format)

        worksheet.merge_range("B7:C9", matkul , text_format)
        worksheet.merge_range("D7:E9", rps_data["kode_matkul"] , text_format)
        worksheet.merge_range("F7:F9", rps_data["rumpun"], text_format)
        worksheet.merge_range("G7:I9", str(int(rps_data["bobot_sks"])), text_format)
        worksheet.merge_range("J7:J9", rps_data["semester"], text_format)
        today = datetime.now()
        worksheet.merge_range("K7:L9", today, date_format)

        # Otorisasi
        worksheet.merge_range("B10:C10", "OTORISASI / PENGESAHAN", title_format)
        worksheet.merge_range("D10:F10", "Dosen Pengembang RPS", title_format)    
        worksheet.merge_range("G10:I10", "Koordinator Mata Kuliah", title_format)    
        worksheet.merge_range("J10:L10", "Ketua Program Studi", title_format)

        worksheet.set_row(10, 110)  # baris 11

        worksheet.merge_range("B11:C11", "OTORISASI / PENGESAHAN", text_otorisasi_format)
        worksheet.merge_range("D11:F11", matkul_data["team_teaching"][0], text_otorisasi_format)    
        worksheet.merge_range("G11:I11", "I Made Adi Bhaskara, S.Kom., M.T.", text_otorisasi_format)    
        worksheet.merge_range("J11:L11", "Ir. I Made Surya Kumara, S.T., M.Sc.", text_otorisasi_format)

        # Placeholder CPL, CPMK, etc.
        cpl_start_row = 12
        worksheet.merge_range(f'C{cpl_start_row}:L{cpl_start_row}', "CPL-PRODI yang dibebankan pada MK", title_cpl_format)
        for i in range(len(cpl_cpmk_sub["cpl_kode"])):
            worksheet.write(f'C{cpl_start_row+1+i}', cpl_cpmk_sub["cpl_kode"][i], text_cpl_format)
            worksheet.merge_range(f'D{cpl_start_row+1+i}:L{cpl_start_row+1+i}', cpl_cpmk_sub["cpl_desc"][i], text_cpl_format)

        cpmk_start_row = cpl_start_row+1+len(cpl_cpmk_sub["cpl_kode"])
        worksheet.merge_range(f'C{cpmk_start_row}:K{cpmk_start_row}', "Capaian Pembelajaran Mata Kuliah (CPMK)", title_cpl_format)
        worksheet.write(f'L{cpmk_start_row}', "Bobot (%)", title_cpl_format)
        for i in range(len(cpl_cpmk_sub["cpmk_kode"])):
            worksheet.write(f'C{cpmk_start_row+1+i}', cpl_cpmk_sub["cpmk_kode"][i], text_cpl_format)
            worksheet.merge_range(f'D{cpmk_start_row+1+i}:K{cpmk_start_row+1+i}', cpl_cpmk_sub["cpmk_desc"][i], text_cpl_format)
            worksheet.write(f'L{cpmk_start_row+1+i}', int(matkul_data["bobot_per_cpmk"][i]), text_cpl_format)

        subcpmk_start_row = cpmk_start_row+1+len(cpl_cpmk_sub["cpmk_kode"])
        worksheet.merge_range(f'C{subcpmk_start_row}:L{subcpmk_start_row}', "Kemampuan akhir tiap tahapan belajar (Sub-CPMK)", title_cpl_format)
        for i in range(len(cpl_cpmk_sub["subcpmk_kode"])):
            worksheet.write(f'C{subcpmk_start_row+1+i}', cpl_cpmk_sub["subcpmk_kode"][i], text_cpl_format)
            worksheet.merge_range(f'D{subcpmk_start_row+1+i}:L{subcpmk_start_row+1+i}', cpl_cpmk_sub["subcpmk_desc"][i], text_cpl_format)

        korelasi_start_row = subcpmk_start_row+1+len(cpl_cpmk_sub["subcpmk_kode"])
        worksheet.merge_range(f'C{korelasi_start_row}:L{korelasi_start_row}', "Korelasi CPL terhadap Sub CPMK", title_cpl_format)

        
        for row in range(korelasi_start_row+1, korelasi_start_row + 2 + len(cpl_cpmk_sub["subcpmk_kode"]) + 1):
            for col in range(2, 12):  # D=3, L=11 (0-based index)
                worksheet.write(row-1, col, "", text_cpl_format)
        
        # baris terakhir untuk total
        total_row = korelasi_start_row + len(cpl_cpmk_sub["subcpmk_kode"]) + 1
        worksheet.write(total_row, 2, "Total", title_korelasi_format)

        for i in range(len(cpl_cpmk_sub["subcpmk_kode"])):
            worksheet.write(f'C{korelasi_start_row+2+i}', cpl_cpmk_sub["subcpmk_kode"][i], title_korelasi_format)

        start_cpl_col = 3
        end_cpl_col = start_cpl_col + len(cpl_cpmk_sub["cpl_kode"])
        for col in range(start_cpl_col,end_cpl_col):
            worksheet.write(korelasi_start_row, col, cpl_cpmk_sub["cpl_kode"][col-3], title_korelasi_format)

        # # Isi bobot sesuai CPL
        # cpl_col_map = {kode: start_cpl_col + idx for idx, kode in enumerate(cpl_cpmk_sub["cpl_kode"])}
        # for i, sub in enumerate(cpl_cpmk_sub["subcpmk_kode"]):
        #     row = korelasi_start_row + 1 + i
        #     kode = matkul_data["cpl_bobot"][i]
        #     bobot = int(float(matkul_data["total_bobot"][i])/100)
        #     if kode in cpl_col_map:
        #         col = cpl_col_map[kode]
        #         worksheet.write(row, col, bobot, percent_format)

        # === Tambahkan SUM total di baris "Total" ===
        # for col in range(start_cpl_col, end_cpl_col):
        #     col_letter = xlsxwriter.utility.xl_col_to_name(col)  # ubah index ke huruf Excel
        #     start_row = korelasi_start_row + 1
        #     end_row = korelasi_start_row + len(cpl_cpmk_sub["subcpmk_kode"])
        #     formula = f"=SUM({col_letter}{start_row+1}:{col_letter}{end_row+1})"
        #     worksheet.write_formula(total_row, col, formula, percent_format_bold)

        # --- isi bobot sesuai CPL (dan akumulasikan totals per kolom) ---
        cpl_col_map = {kode: start_cpl_col + idx for idx, kode in enumerate(cpl_cpmk_sub["cpl_kode"])}

        # inisialisasi totals per kolom (numeric col index)
        totals = {col: 0.0 for col in range(start_cpl_col, end_cpl_col)}

        for i, sub in enumerate(cpl_cpmk_sub["subcpmk_kode"]):
            excel_row = korelasi_start_row + 1 + i               # Excel row number (1-based)
            kode = matkul_data["cpl_bobot"][i]
            # bobot sebagai fraction (mis. 25 -> 0.25) untuk format persen
            try:
                bobot = float(matkul_data["total_bobot"][i]) / 100.0
            except Exception:
                bobot = 0.0

            if kode in cpl_col_map:
                col = cpl_col_map[kode]                          # numeric col (0-based as used before)
                col_letter = xlsxwriter.utility.xl_col_to_name(col)  # convert to letter, ex. 3 -> 'D'
                # tulis nilai persen ke sel, pakai format percent_format
                worksheet.write(f"{col_letter}{excel_row+1}", bobot, percent_format)
                # akumulasikan total per kolom
                totals[col] += bobot

        # === tulis total (sebagai angka persen) di baris "Total" tanpa formula ===
        total_row = korelasi_start_row + len(cpl_cpmk_sub["subcpmk_kode"]) + 2  # baris Excel tempat "Total"
        total_per_cpl = []
        for col in range(start_cpl_col, end_cpl_col):
            col_letter = xlsxwriter.utility.xl_col_to_name(col)
            total_value = totals.get(col, 0.0)
            total_per_cpl.append(total_value)
            worksheet.write(f"{col_letter}{total_row}", total_value, percent_format_bold)

        worksheet.merge_range(f"B{cpl_start_row}:B{total_row}", "Capaian Pembelajaran", title_korelasi_format)

        desc_start_row = total_row + 1
        worksheet.set_row(desc_start_row-1, 110)  # baris 11
        worksheet.write(f"B{desc_start_row}", "Deskripsi Singkat MK", title_korelasi_format)
        
        materi_str = ", ".join(matkul_data["materi_non_uts_uas"])
        description_matkul = (
            f"Mata kuliah {matkul} membahas konsep teoritis, metode, "
            f"dan implementasi mengenai materi seperti {materi_str}."
        )
        worksheet.merge_range(f'C{desc_start_row}:L{desc_start_row}', description_matkul, text_cpl_format)

        bahan_start_row = desc_start_row + 1
        bahan_end_row = bahan_start_row + len(matkul_data["materi_non_uts_uas_numbered"]) - 1
        worksheet.merge_range(f"B{bahan_start_row}:B{bahan_end_row}", "Bahan Kajian/Materi Pembelajaran", title_korelasi_format)
        for i in range(len(matkul_data["materi_non_uts_uas_numbered"])):
            worksheet.merge_range(f'C{desc_start_row+1+i}:L{desc_start_row+1+i}', matkul_data["materi_non_uts_uas_numbered"][i], text_cpl_format)

        pustaka_start_row = bahan_end_row + 1
        pustaka_utama_end_row = pustaka_start_row + len(matkul_data["pustaka_utama"])
        pustaka_pendukung_end_row = pustaka_utama_end_row + len(matkul_data["pustaka_pendukung"]) + 1
        worksheet.merge_range(f"B{pustaka_start_row}:B{pustaka_pendukung_end_row}", "Pustaka", title_korelasi_format)
        worksheet.merge_range(f"C{pustaka_start_row}:L{pustaka_start_row}", "Utama:", title_cpl_format)
        worksheet.merge_range(f"C{pustaka_utama_end_row+1}:L{pustaka_utama_end_row+1}", "Pendukung:", title_cpl_format)
        
        for i in range(len(matkul_data["pustaka_utama"])):
            worksheet.merge_range(f'C{pustaka_start_row+1+i}:L{pustaka_start_row+1+i}', matkul_data["pustaka_utama"][i], text_cpl_format)
        
        for i in range(len(matkul_data["pustaka_pendukung"])):
            worksheet.merge_range(f'C{pustaka_utama_end_row+2+i}:L{pustaka_utama_end_row+2+i}', matkul_data["pustaka_pendukung"][i], text_cpl_format)
        
        dosen_start_row = pustaka_pendukung_end_row + 1
        dosen_end_row = dosen_start_row + len(matkul_data["team_teaching"]) - 1
        worksheet.merge_range(f"B{dosen_start_row}:B{dosen_end_row}", "Dosen Pengampu", title_korelasi_format)
        for i in range(len(matkul_data["team_teaching"])):
            worksheet.merge_range(f'C{dosen_start_row+i}:L{dosen_start_row+i}', matkul_data["team_teaching"][i], text_cpl_format)

        syarat_start_row = dosen_end_row + 1
        dosen_end_row = syarat_start_row + len(matkul_data["matkul_syarat"]) - 1
        if dosen_end_row == syarat_start_row:
            worksheet.write(f"B{syarat_start_row}", "Matakuliah Syarat", title_korelasi_format)
        else:
            worksheet.merge_range(f"B{syarat_start_row}:B{dosen_end_row}", "Matakuliah Syarat", title_korelasi_format)
        for i in range(len(matkul_data["matkul_syarat"])):
            worksheet.merge_range(f'C{syarat_start_row+i}:L{syarat_start_row+i}', matkul_data["matkul_syarat"][i], text_cpl_format)

        # Pertemuan Mingguan 
        # Header
        mingguan_start_row = dosen_end_row + 1 
        worksheet.merge_range(f"B{mingguan_start_row}:B{mingguan_start_row+2}", "Mg Ke-", title_format)
        worksheet.write(f"B{mingguan_start_row+3}", "(1)", title_format)

        worksheet.merge_range(f"C{mingguan_start_row}:C{mingguan_start_row+2}", "Kemampuan akhir tiap tahapan belajar (Sub-CPMK)", title_format)
        worksheet.write(f"C{mingguan_start_row+3}", "(2)", title_format)

        worksheet.merge_range(f"D{mingguan_start_row}:G{mingguan_start_row}", "Penilaian", title_format)
        worksheet.merge_range(f"D{mingguan_start_row+1}:E{mingguan_start_row+2}", "Indikator", title_format)
        worksheet.merge_range(f"F{mingguan_start_row+1}:G{mingguan_start_row+2}", "Teknik dan Instrumen Penilaian", title_format)
        worksheet.merge_range(f"D{mingguan_start_row+3}:E{mingguan_start_row+3}", "(3)", title_format)
        worksheet.merge_range(f"F{mingguan_start_row+3}:G{mingguan_start_row+3}", "(4)", title_format)

        worksheet.merge_range(f"H{mingguan_start_row}:I{mingguan_start_row+1}", "Strategi Pembelajaran dan Metode Pembelajaran [Estimasi Waktu]", title_format)
        worksheet.write(f"H{mingguan_start_row+2}", "Luring (offline)", title_format)
        worksheet.write(f"I{mingguan_start_row+2}", "Daring (online)", title_format)
        worksheet.write(f"H{mingguan_start_row+3}", "(5)", title_format)
        worksheet.write(f"I{mingguan_start_row+3}", "(6)", title_format)

        worksheet.merge_range(f"J{mingguan_start_row}:K{mingguan_start_row+2}", "Materi Pembelajaran dan Daftar Referensi \n[Pustaka]", title_format)
        worksheet.merge_range(f"J{mingguan_start_row+3}:K{mingguan_start_row+3}", "(7)", title_format)

        worksheet.merge_range(f"L{mingguan_start_row}:L{mingguan_start_row+2}", "Bobot Penilaian (%)", title_format)
        worksheet.write(f"L{mingguan_start_row+3}", "(8)", title_format)

        # Body RPS
        mingguan_body_start_row = mingguan_start_row + 4
        len_mingguan = len(matkul_data["minggu_ke"])

        weekly_subcpmk_desc = []

        for i in range(len_mingguan):
            worksheet.write(f'B{mingguan_body_start_row+i}', int(float(matkul_data["minggu_ke"][i])), text_cpl_format)

            # === cari deskripsi subcpmk ===
            subcpmk_kode = matkul_data["subcpmk_weekly"][i]
            subcpmk_desc = ""
            if "subcpmk_kode" in cpl_cpmk_sub and "subcpmk_desc" in cpl_cpmk_sub:
                if subcpmk_kode in cpl_cpmk_sub["subcpmk_kode"]:
                    idx = cpl_cpmk_sub["subcpmk_kode"].index(subcpmk_kode)
                    subcpmk_desc = cpl_cpmk_sub["subcpmk_desc"][idx]

            weekly_subcpmk_desc.append(subcpmk_desc)
            worksheet.write(
                f'C{mingguan_body_start_row+i}',
                f'{subcpmk_desc} ({subcpmk_kode}) ',
                text_cpl_format
            )

            indikator_text = matkul_data["indikator_numbered"][i]

            # === Cek apakah ini baris evaluasi ===
            if "Evaluasi UTS" in indikator_text or "Evaluasi UAS" in indikator_text:
                # Merge dari C sampai K, isi dengan teks evaluasi
                worksheet.merge_range(
                    f'D{mingguan_body_start_row+i}:K{mingguan_body_start_row+i}',
                    indikator_text,
                    title_format
                )
                # Tetap isi bobot di L
                worksheet.write(f'L{mingguan_body_start_row+i}', int(float(matkul_data["bobot"][i])), text_cpl_format)
                continue  # skip ke iterasi berikutnya

            worksheet.merge_range(f'D{mingguan_body_start_row+i}:E{mingguan_body_start_row+i}', indikator_text, text_cpl_format)
            worksheet.merge_range(f'F{mingguan_body_start_row+i}:G{mingguan_body_start_row+i}', matkul_data["kriteria_numbered"][i], text_cpl_format)

            if "Tugas" in matkul_data["kriteria_numbered"][i]:
                worksheet.write(f'H{mingguan_body_start_row+i}', f"Ekspository dan diskusi [TM : {rps_data['bobot_sks']}x50'] Task Based Learning [TB : {rps_data['bobot_sks']}x50']", text_cpl_format)
            else:
                worksheet.write(f'H{mingguan_body_start_row+i}', f"Ekspository dan diskusi [TM : {rps_data['bobot_sks']}x50']", text_cpl_format)

            worksheet.write(f'I{mingguan_body_start_row+i}', f"Link materi [BM : {rps_data['bobot_sks']}x50']", text_cpl_format)

            worksheet.merge_range(
                f'J{mingguan_body_start_row+i}:K{mingguan_body_start_row+i}',
                f'{matkul_data["materi_weekly_numbered"][i]} \n[{matkul_data["pustaka_weekly"][i]}]',
                text_cpl_format
            )

            worksheet.write(f'L{mingguan_body_start_row+i}', float(matkul_data["bobot"][i]), text_cpl_format)

        blueprint_start_row = mingguan_body_start_row + len_mingguan + 2

        worksheet.merge_range(f"B{blueprint_start_row}:L{blueprint_start_row}", "BLUE PRINT PENILAIAN ATAU RENCANA ASESMEN DAN EVALUASI (RAE)", title_format)
        
        start_col_green = 1  # B
        end2_col_green = 11
        end_col_green = 2 + len(cpl_cpmk_sub["cpl_kode"])  # geser ke kanan sesuai jumlah CPL

        start_green_cell = xl_rowcol_to_cell(blueprint_start_row, start_col_green)
        end_green_cell = xl_rowcol_to_cell(blueprint_start_row, end_col_green)
        start2_green_cell = xl_rowcol_to_cell(blueprint_start_row, end_col_green+1)
        end2_green_cell = xl_rowcol_to_cell(blueprint_start_row, end2_col_green)

        worksheet.merge_range(f'{start_green_cell}:{end_green_cell}', "KRITERIA PENILAIAN", title_format_green)
        worksheet.merge_range(f'{start2_green_cell}:{end2_green_cell}', "RANGE NILAI", title_format_green)

        worksheet.merge_range(f'B{blueprint_start_row+2}:B{blueprint_start_row+3}', "Bobot", title_format)
        worksheet.merge_range(f'C{blueprint_start_row+2}:C{blueprint_start_row+3}', "Teknik dan Penilaian", title_format)

        for col in range(start_cpl_col,end_cpl_col):
            worksheet.write(blueprint_start_row + 1, col, cpl_cpmk_sub["cpl_kode"][col-3], title_format)
            worksheet.write(blueprint_start_row + 2, col, total_per_cpl[col-3], percent_format_bold_fill)
        
        start_asm_cell = xl_rowcol_to_cell(blueprint_start_row+1, end_col_green+1)
        end_asm_cell = xl_rowcol_to_cell(blueprint_start_row+2, end_col_green+1)

        worksheet.merge_range(f'{start_asm_cell}:{end_asm_cell}', "Bobot Asesmen", title_format)

        start_range_cell = xl_rowcol_to_cell(blueprint_start_row+1, end_col_green+2)
        end_range_cell = xl_rowcol_to_cell(blueprint_start_row+7, 11)    

        worksheet.merge_range(f'{start_range_cell}:{end_range_cell}', "Nilai akhir diatas dikonversikan kedalam huruf mutu menggunakan kriteria penilaian sebagai berikut:\nRENTANGAN NILAI :\n85.00 - 100.00 : A (UNGGUL - LULUS)                \n75.00 - 84.99   : AB (BAIK SEKALI - LULUS)              \n70.00 - 74.99   : B (BAIK - LULUS)             \n60.00 - 69.99   : BC (CUKUP BAIK - TIDAK LULUS)\n55.00 - 59.99   : C (CUKUP - TIDAK LULUS)\n50.00 - 54.99   : CD (KURANG - TIDAK LULUS)\n44.00 - 49.99   : D (KURANG SEKALI - TIDAK LULUS)\n0.00 - 43.99     : E (GAGAL - TIDAK LULUS)", text_cpl_format)

        # tentukan start row/col dan end row/col
        start_row = blueprint_start_row + 3
        start_col = 3
        end_row = blueprint_start_row + 7
        end_col = end_col_green + 1

        # isi semua cell di range dengan border
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                worksheet.write(row, col, "", text_format)
            for i in range(len(cpl_cpmk_sub["subcpmk_kode"])):
                worksheet.write(f'B{blueprint_start_row+4+i}', cpl_cpmk_sub["subcpmk_kode"][i], text_format)

        kriteria_per_subcpmk = []
        rubrik_per_subcpmk = []

        for kode in cpl_cpmk_sub["subcpmk_kode"]:
            # cari semua kriteria untuk subcpmk ini
            related_kriteria = [
                matkul_data["kriteria"][i]
                for i, wk in enumerate(matkul_data["subcpmk_weekly"])
                if wk == kode
            ]

            # cek apakah ada "Tugas" 
            if any("Tugas" in k for k in related_kriteria): 
                kriteria_per_subcpmk.append("Ekspository dan diskusi (Oral Assessment), Multiple Choice Questions (MCQ) dan Short Answer Questions (SAQ)") 
            else: 
                kriteria_per_subcpmk.append("Kuis, diskusi, dan wawancara pemahaman (Oral Assessment)")
            
            # kumpulkan semua teks di dalam [ ... ]
            rubrik_items = []
            if related_kriteria:
                for k in related_kriteria:
                    matches = re.findall(r"\[(.*?)\]", k)
                    rubrik_items.extend(matches)
            rubrik_per_subcpmk.append(", ".join(rubrik_items) if rubrik_items else "")

        for i in range(len(cpl_cpmk_sub["subcpmk_kode"])): 
            worksheet.write(f'C{blueprint_start_row+4+i}', kriteria_per_subcpmk[i], text_format)

        for i, sub in enumerate(cpl_cpmk_sub["subcpmk_kode"]):
            excel_row = blueprint_start_row + 3 + i               # Excel row number (1-based)
            kode = matkul_data["cpl_bobot"][i]
            # bobot sebagai fraction (mis. 25 -> 0.25) untuk format persen
            try:
                bobot = int(float(matkul_data["total_bobot"][i]))
            except Exception:
                bobot = 0

            if kode in cpl_col_map:
                col = cpl_col_map[kode]                          # numeric col (0-based as used before)
                col_letter = xlsxwriter.utility.xl_col_to_name(col)  # convert to letter, ex. 3 -> 'D'
                # tulis nilai persen ke sel, pakai format percent_format
                worksheet.write(f"{col_letter}{excel_row+1}", f'Nilai x {bobot}% \n({rubrik_per_subcpmk[i]})', percent_format)

        for i in range(len(cpl_cpmk_sub["subcpmk_kode"])):
            worksheet.write(blueprint_start_row+3+i, end_col_green+1, matkul_data["total_bobot"][i], text_format)

        last_rps_start_row = blueprint_start_row + len(cpl_cpmk_sub["subcpmk_kode"]) + 4
        worksheet.merge_range(f'B{last_rps_start_row}:C{last_rps_start_row}', "JUMLAH", title_korelasi_format)
        worksheet.merge_range(f'B{last_rps_start_row+1}:C{last_rps_start_row+1}', "NILAI MATA KULIAH", title_korelasi_format)
        worksheet.merge_range(f'B{last_rps_start_row+2}:C{last_rps_start_row+2}', "NILAI CPL", title_korelasi_format)

        # buat label AA, BB, CC, ... sesuai jumlah CPL
        labels = [chr(65 + i) * 2 for i in range(len(cpl_cpmk_sub["cpl_kode"]))]  # A=65 di ASCII

        # tulis label di row last_rps_start_row
        for idx, col in enumerate(range(start_cpl_col, end_cpl_col)):
            worksheet.write(last_rps_start_row-1, col, labels[idx], title_korelasi_format)
            worksheet.write(
                last_rps_start_row + 1,
                col,
                f"{labels[idx]}/{total_per_cpl[idx]} x100",
                title_korelasi_format,
            )

        # merge semua kolom di row last_rps_start_row untuk teks gabungan
        merged_start = xlsxwriter.utility.xl_col_to_name(start_cpl_col)
        merged_end = xlsxwriter.utility.xl_col_to_name(end_cpl_col - 1)
        worksheet.merge_range(
            f"{merged_start}{last_rps_start_row+1}:{merged_end}{last_rps_start_row+1}",
            " + ".join(labels),
            title_korelasi_format,
        )

        ######################## RPM ######################
        def write_rpm_template(rpm_sheet_name, judul_kriteria, subcpmk_rpm, indikator_numbered_rpm, minggu_rpm, bobot_rpm):
            worksheet_rpm = workbook.add_worksheet(rpm_sheet_name)
            # Header
            # Tambahkan logo
            # Set tinggi baris header (sedikit lebih tinggi dari normal)
            # Atur ukuran kolom
            worksheet_rpm.set_column("A:A", 5)        # Kolom A kecil
            worksheet_rpm.set_column("B:B", 18)       # Kolom B - L agak besar (2x normal)
            worksheet_rpm.set_column("C:J", 18)       # Kolom B - L agak besar (2x normal)
            worksheet_rpm.set_row(1, 22)  # baris 2
            worksheet_rpm.set_row(2, 22)  # baris 3
            worksheet_rpm.set_row(3, 22)  # baris 4
            worksheet_rpm.set_row(4, 26)  # baris 5
            worksheet_rpm.merge_range("B2:B5", "", header_medium)
            worksheet_rpm.insert_image("B2", "data/logo.png", {
                "x_scale": 1.5,  # perkecil jika perlu
                "y_scale": 1.5,
                "x_offset": 10,  # sedikit geser biar rapi
                "y_offset": 2,
            })
            
            worksheet_rpm.merge_range("C2:H2", "UNIVERSITAS WARMADEWA", header_medium)
            worksheet_rpm.merge_range("C3:H3", "FAKULTAS TEKNIK DAN PERENCANAAN", header_medium)
            worksheet_rpm.merge_range("C4:H4", "PROGRAM STUDI TEKNIK KOMPUTER", header_medium)
            worksheet_rpm.merge_range("C5:H5", "RENCANA PENUGASAN MAHASISWA", header_big)

            kode_dokumen_rpm = f'FTP-TKOM-RPM-{rps_data["kode_matkul"]}-{tahun}'
            worksheet_rpm.merge_range("I2:J3", "Kode Dokumen", header_small)
            worksheet_rpm.merge_range("I4:J5", str(kode_dokumen_rpm), header_small)

            worksheet_rpm.merge_range("B6:C6", "MATA KULIAH (MK)", title_cpl_format)
            worksheet_rpm.merge_range("D6:J6", matkul, text_cpl_format)

            worksheet_rpm.merge_range("B7:C7", "KODE", title_cpl_format)
            worksheet_rpm.merge_range("D7:E7", rps_data["kode_matkul"], text_cpl_format)
            worksheet_rpm.write("F7", "SKS", title_cpl_format)
            worksheet_rpm.write("G7", str(int(rps_data["bobot_sks"])), text_cpl_format)
            worksheet_rpm.write("H7", "SEMESTER", title_cpl_format)
            worksheet_rpm.merge_range("I7:J7", rps_data["semester"], text_cpl_format)

            worksheet_rpm.merge_range("B8:C11", "DOSEN PENGAMPU", title_cpl_format)
            # Buat 4 baris kosong dulu
            for i in range(4):
                worksheet_rpm.merge_range(f"D{8+i}:J{8+i}", "", text_cpl_format)

            # Isi nama dosen sesuai jumlah
            if len(matkul_data["team_teaching"]) < 5:
                for i in range(len(matkul_data["team_teaching"])):                
                    worksheet_rpm.write(f"D{8+i}", matkul_data["team_teaching"][i], text_cpl_format)
            else:
                worksheet_rpm.write("D8", matkul_data["team_teaching"][0], text_cpl_format)

            worksheet_rpm.merge_range("B12:F12", "BENTUK TUGAS", title_cpl_format)
            worksheet_rpm.merge_range("B13:F13", "Penugasan Individu", text_cpl_format)
            worksheet_rpm.merge_range("G12:J12", "WAKTU PENGERJAAN TUGAS", title_cpl_format)
            worksheet_rpm.merge_range("G13:J13", f'Minggu ke-{int(float(minggu_rpm))}', text_cpl_format)

            worksheet_rpm.merge_range("B14:J14", "JUDUL TUGAS", title_cpl_format)
            worksheet_rpm.merge_range("B15:J15", judul_kriteria, text_cpl_format)

            worksheet_rpm.merge_range("B16:J16", "SUB CAPAIAN PEMBELAJARAN MATA KULIAH", title_cpl_format)
            worksheet_rpm.merge_range("B17:J17", subcpmk_rpm, text_cpl_format)

            worksheet_rpm.merge_range("B18:J18", "DESKRIPSI TUGAS", title_cpl_format)
            worksheet_rpm.merge_range("B19:J19", indikator_numbered_rpm, text_cpl_format)

            worksheet_rpm.merge_range("B20:J20", "METODE PENGERJAAN TUGAS", title_cpl_format)
            worksheet_rpm.merge_range("B21:J21", "Mahasiswa menjawab soal yang diberikan pada saat perkuliahan", text_cpl_format)
            worksheet_rpm.merge_range("B22:J22", "BENTUK DAN FORMAT LUARAN", title_cpl_format)
            worksheet_rpm.write("B23", "a. Obyek Garapan", title_korelasi_format)
            worksheet_rpm.merge_range("C23:J23", "Daftar soal", text_cpl_format)
            worksheet_rpm.write("B24", "b. Bentuk Luaran", title_korelasi_format)
            worksheet_rpm.merge_range("C24:J24", "Penjelasan dan analisis", text_cpl_format)
            
            worksheet_rpm.merge_range("B25:J25", "INDIKATOR, KRITERIA, dan BOBOT PENILAIAN", title_cpl_format)
            worksheet_rpm.merge_range("B26:J26", f'Indikator: {indikator_numbered_rpm}', text_cpl_format)
            worksheet_rpm.merge_range("B27:J27", "", text_cpl_format)
            worksheet_rpm.merge_range("B28:J28", f'Bobot Penilaian : {bobot_rpm} % dari total 100% penilaian mata kuliah', text_cpl_format)
            worksheet_rpm.merge_range("B29:J29", "Kriteria Penilaian: Terlampir", text_cpl_format)

            worksheet_rpm.merge_range("B30:J30", "JADWAL PELAKSANAAN", title_cpl_format)
            worksheet_rpm.merge_range("B31:J31", f'Minggu ke-{int(float(minggu_rpm))}', text_cpl_format)

            worksheet_rpm.merge_range("B32:J32", "LAIN-LAIN", title_cpl_format)
            worksheet_rpm.merge_range("B33:J33", "-", text_cpl_format)

            worksheet_rpm.merge_range("B34:J34", "REFERENSI", title_cpl_format)
            for i in range(len(matkul_data["pustaka_utama"])):
                worksheet_rpm.merge_range(f'B{35+i}:J{35+i}', matkul_data["pustaka_utama"][i], text_cpl_format)
            
            for i in range(len(matkul_data["pustaka_pendukung"])):
                worksheet_rpm.merge_range(f'B{35+len(matkul_data["pustaka_utama"])+i}:J{35+len(matkul_data["pustaka_utama"])+i}', matkul_data["pustaka_pendukung"][i], text_cpl_format)
   
        # --- Variabel kontrol ---
        def to_number(value):
            try:
                return float(value)
            except (ValueError, TypeError):
                return 0
        
        rpm_index = 1
        tugas_count = 0
        kuis_count = 0
        has_uts = False
        has_uas = False

        # --- Variabel agregasi untuk UTS & UAS ---
        uts_minggu = None
        uts_bobot_total = 0
        uts_indikator = None

        uas_minggu = None
        uas_bobot_total = 0
        uas_indikator = None

        # --- Loop semua kriteria ---
        for i in range(len(matkul_data["kriteria_numbered"])):
            kriteria = matkul_data["kriteria_numbered"][i]

            if "Tugas" in kriteria:
                tugas_count += 1
                sheet_name = f"RPM{rpm_index} (Tugas {tugas_count})"
                write_rpm_template(
                    sheet_name,
                    kriteria,
                    weekly_subcpmk_desc[i],
                    matkul_data["indikator_numbered"][i],
                    matkul_data["minggu_ke"][i],
                    matkul_data["bobot"][i],
                )
                rpm_index += 1

            elif "Kuis" in kriteria:
                kuis_count += 1
                sheet_name = f"RPM{rpm_index} (Kuis {kuis_count})"
                write_rpm_template(
                    sheet_name,
                    kriteria,
                    weekly_subcpmk_desc[i],
                    matkul_data["indikator_numbered"][i],
                    matkul_data["minggu_ke"][i],
                    matkul_data["bobot"][i],
                )
                rpm_index += 1

            elif "Evaluasi UTS" in kriteria:
                has_uts = True
                uts_bobot_total += to_number(matkul_data["bobot"][i])
                if uts_minggu is None:
                    uts_minggu = matkul_data["minggu_ke"][i]
                if uts_indikator is None:
                    uts_indikator = matkul_data["indikator_numbered"][i]

            elif "Evaluasi UAS" in kriteria:
                has_uas = True
                uas_bobot_total += to_number(matkul_data["bobot"][i])
                if uas_minggu is None:
                    uas_minggu = matkul_data["minggu_ke"][i]
                if uas_indikator is None:
                    uas_indikator = matkul_data["indikator_numbered"][i]

        # --- Tambahkan sheet UTS kalau ada ---
        if has_uts:
            sheet_name = f"RPM{rpm_index} (Evaluasi UTS)"
            write_rpm_template(
                sheet_name,
                "Evaluasi UTS",
                "Evaluasi UTS",
                uts_indikator,
                uts_minggu,
                uts_bobot_total,
            )
            rpm_index += 1

        # --- Tambahkan sheet UAS kalau ada ---
        if has_uas:
            sheet_name = f"RPM{rpm_index} (Evaluasi UAS)"
            write_rpm_template(
                sheet_name,
                "Evaluasi UAS",
                "Evaluasi UAS",
                uas_indikator,
                uas_minggu,
                uas_bobot_total,
            )
            rpm_index += 1

        ############################## RUBRIK ################################
        rubrik_list = [
            ("SP1", "Skala Persepsi", "Rubrik Penilaian Presentasi Lisan Mahasiswa"),
            ("H1", "Holistik", "Rubrik Penilaian Penugasan Mahasiswa"),
            ("H2", "Holistik", "Rubrik Penilaian UTS/UAS Mahasiswa"),
            ("H3", "Holistik", "Rubrik Penilaian Rancangan Proposal Mahasiswa"),
            ("A1", "Analitik", "Rubrik Penilaian Presentasi Makalah Mahasiswa"),
            ("A2", "Analitik", "Rubrik Penilaian Project Based Learning Mahasiswa"),
            ("A3", "Analitik", "Rubrik Penilaian Capstone Project Mahasiswa"),
        ]

        for kode_rubrik, type_rubrik, header_title in rubrik_list:
            subcpmk_rub_key = f"rubrik_{kode_rubrik}_subcpmk"
            cpl_rub_key = f"rubrik_{kode_rubrik}_cpl"

            subcpmk_rub_data = matkul_data[subcpmk_rub_key]
            cpl_rub_data = matkul_data[cpl_rub_key]

            if not subcpmk_rub_data or not cpl_rub_data:
                continue  # skip kalau kosong

            # --- Buat worksheet ---
            sheet_title = f"RUB {kode_rubrik}"
            worksheet_rub = workbook.add_worksheet(sheet_title)

            # Atur ukuran kolom
            worksheet_rub.set_column("A:A", 5)        # Kolom A kecil
            worksheet_rub.set_column("B:I", 25)       # Kolom B - L agak besar (2x normal)
            worksheet_rub.set_row(1, 22)  # baris 2
            worksheet_rub.set_row(2, 22)  # baris 3
            worksheet_rub.set_row(3, 22)  # baris 4
            worksheet_rub.set_row(4, 22)  # baris 5
            worksheet_rub.merge_range("B2:B5", "", header_medium)
            worksheet_rub.insert_image("B2", "data/logo.png", {
                "x_scale": 1.5,  # perkecil jika perlu
                "y_scale": 1.5,
                "x_offset": 10,  # sedikit geser biar rapi
                "y_offset": 2,
            })

            # --- Header utama ---            
            worksheet_rub.merge_range("C2:G2", "UNIVERSITAS WARMADEWA", header_medium)
            worksheet_rub.merge_range("C3:G3", "FAKULTAS TEKNIK DAN PERENCANAAN", header_medium)
            worksheet_rub.merge_range("C4:G4", "PROGRAM STUDI TEKNIK KOMPUTER", header_medium)
            worksheet_rub.merge_range("C5:G5", header_title, header_small)

            kode_dokumen_rub = f'FTP-TKOM-RUB-{rps_data["kode_matkul"]}-{tahun}'
            worksheet_rub.merge_range("H2:I3", "Kode Dokumen", header_small)
            worksheet_rub.merge_range("H4:I5", str(kode_dokumen_rub), header_small)

            # # --- Info MK ---
            worksheet_rub.merge_range("B6:C6", "MATA KULIAH", title_cpl_format)
            worksheet_rub.merge_range("D6:I6", matkul, text_cpl_format)

            worksheet_rub.merge_range("B7:C7", "KODE", title_cpl_format)
            worksheet_rub.merge_range("D7:I7", rps_data["kode_matkul"], text_cpl_format)

            worksheet_rub.merge_range("B8:C11", "DOSEN PENGAMPU", title_cpl_format)
            # Buat 4 baris kosong dulu
            for i in range(4):
                worksheet_rub.merge_range(f"D{8+i}:I{8+i}", "", text_cpl_format)

            # Isi nama dosen sesuai jumlah
            if len(matkul_data["team_teaching"]) < 5:
                for i in range(len(matkul_data["team_teaching"])):                
                    worksheet_rub.write(f"D{8+i}", matkul_data["team_teaching"][i], text_cpl_format)
            else:
                worksheet_rub.write("D8", matkul_data["team_teaching"][0], text_cpl_format)

            worksheet_rub.merge_range("B12:C12", "SEMESTER", title_cpl_format)
            worksheet_rub.merge_range("D12:I12", rps_data["semester"], text_cpl_format)

            worksheet_rub.merge_range("B13:C13", "SKS", title_cpl_format)
            worksheet_rub.merge_range("D13:I13", rps_data["bobot_sks"], text_cpl_format)

            # --- Cari semua kriteria yang ada kode rubrik (misal "SP1") ---
            related_kriteria = [
                k for k in matkul_data["kriteria_numbered"] if kode_rubrik in k
            ]
            tugas_str = "\n ".join(related_kriteria)

            worksheet_rub.merge_range("B14:C14", "Tugas", title_cpl_format)
            worksheet_rub.merge_range("D14:I14", tugas_str, text_cpl_format)

            worksheet_rub.merge_range("B15:C15", "Tipe", title_cpl_format)
            worksheet_rub.merge_range("D15:I15", type_rubrik, text_cpl_format)

            worksheet_rub.merge_range("B16:C16", "Sifat", title_cpl_format)
            worksheet_rub.merge_range("D16:I16", "Individu", text_cpl_format)

            # --- Capaian & deskripsi ---
            worksheet_rub.merge_range("B17:C17", "Capaian", title_cpl_format)

            row = 16
            # Buat 4 baris kosong dulu
            for i in range(len(subcpmk_rub_data)):
                worksheet_rub.merge_range(f"E{17+i}:I{17+i}", "", text_cpl_format)

            for idx, (subcpmk, cpl) in enumerate(zip(subcpmk_rub_data, cpl_rub_data)):
                try:
                    sub_desc = cpl_cpmk_sub["subcpmk_desc"][idx]  # ambil berdasarkan urutan
                except (IndexError, TypeError):
                    sub_desc = ""
                worksheet_rub.write(row, 3, subcpmk, title_korelasi_format)
                worksheet_rub.write(row, 4, sub_desc, text_cpl_format)
                row += 1
        
        workbook.close()
        output.seek(0)

        return send_file(
            output,
            as_attachment=True,
            download_name=f"RPS_RPM_RUB_{matkul}_{tahun}.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    except Exception as e:
        logger.error(f"Error generating Excel file: {e}")
        abort(500, description=f"Terjadi kesalahan saat membuat file Excel: {str(e)}")

@app.route("/download-template")
def download_template():
    return send_from_directory(
        "data", 
        "Template Rubrik.xlsx", 
        as_attachment=True
    )

# Add error handlers
@app.errorhandler(400)
def bad_request(error):
    return render_template('error.html', 
                        error_code=400, 
                        error_message=error.description), 400

@app.errorhandler(404)
def not_found(error):
    return render_template('error.html',
                        error_code=404, 
                        error_message=error.description), 404

@app.errorhandler(500)
def internal_error(error):
    return render_template('error.html', 
                        error_code=500, 
                        error_message=error.description), 500

if __name__ == "__main__":
    app.run(debug=True)